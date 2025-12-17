# tasks/admin.py
from django.contrib import admin
from django.utils.translation import gettext_lazy as _
from django.forms import TextInput, Textarea
from django.db import models
from django.urls import path, reverse
from django.shortcuts import render, get_object_or_404
from django.utils.html import format_html
from django.http import HttpResponse
from django.utils import timezone
from datetime import timedelta
from nested_admin import NestedModelAdmin, NestedStackedInline, NestedTabularInline
from .models import (
    Task, TaskStatus, TaskType, SurveyQuestion, 
    SurveyQuestionChoice, SurveyAnswer, PhotoReport, PhotoReportItem,
    SurveyAnswerPhoto, SurveyAnswerGroupReadStatus
)
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Import the new API functions
from .views import getGroupedAnswers, markAsRead, autocomplete_clients, autocomplete_tasks

class SurveyQuestionChoiceInline(NestedTabularInline):
    """Inline choices for survey questions."""
    model = SurveyQuestionChoice
    extra = 3
    verbose_name = _('–í–∞—Ä–∏–∞–Ω—Ç –æ—Ç–≤–µ—Ç–∞')
    verbose_name_plural = _('–í–∞—Ä–∏–∞–Ω—Ç—ã –æ—Ç–≤–µ—Ç–æ–≤')

class SurveyQuestionInline(NestedStackedInline):
    """Inline questions for survey tasks."""
    model = SurveyQuestion
    extra = 1
    inlines = [SurveyQuestionChoiceInline]
    verbose_name = _('–í–æ–ø—Ä–æ—Å')
    verbose_name_plural = _('–í–æ–ø—Ä–æ—Å—ã')
    formfield_overrides = {
        models.CharField: {'widget': TextInput(attrs={'size': '80'})},
        models.TextField: {'widget': Textarea(attrs={'rows': 3, 'cols': 80})},
    }

# SurveyQuestionChoice is hidden from admin as per requirements
# @admin.register(SurveyQuestionChoice)
# class SurveyQuestionChoiceAdmin(admin.ModelAdmin):
#     list_display = ('question', 'choice_text', 'order')
#     list_filter = ('question__task', 'question')
#     search_fields = ('choice_text', 'question__question_text')
#     ordering = ('question', 'order')

@admin.register(Task)
class TaskAdmin(NestedModelAdmin):
    list_display = ('title', 'task_type', 'status', 'is_active', 
                   'assigned_to', 'client', 'created_by', 'created_at',
                   'get_completion_info')
    list_filter = ('task_type', 'status', 'is_active', 'assigned_to', 'client', 'created_by')
    search_fields = ('title', 'description')
    list_per_page = 20
    fieldsets = (
        (_('–û—Å–Ω–æ–≤–Ω–∞—è –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏—è'), {
            'fields': ('title', 'description', 'task_type', 'status', 'is_active')
        }),
        (_('–ù–∞–∑–Ω–∞—á–µ–Ω–∏–µ'), {
            'fields': ('assigned_to', 'client', 'created_by'),
            'classes': ('wide',)
        }),
        (_('–ü–ª–∞–Ω –≤—ã–ø–æ–ª–Ω–µ–Ω–∏—è'), {
            'fields': ('target_count', 'current_count'),
            'classes': ('collapse',)
        }),
        (_('–î–æ–ø–æ–ª–Ω–∏—Ç–µ–ª—å–Ω–æ'), {
            'fields': ('moderator_comment',),
            'classes': ('collapse',)
        }),
    )
    
    def get_inlines(self, request, obj=None):
        if obj and obj.task_type == TaskType.SURVEY:
            return [SurveyQuestionInline]
        elif obj and obj.task_type in [TaskType.EQUIPMENT_PHOTO, TaskType.SIMPLE_PHOTO]:
            return []
        return []
    
    def get_queryset(self, request):
        return super().get_queryset(request).select_related(
            'assigned_to', 'client', 'created_by'
        )
    
    def get_completion_info(self, obj):
        if obj.task_type == TaskType.SURVEY:
            percentage = obj.get_completion_percentage()
            return format_html(
                '{} / {} ({}%)<br><a href="{}" class="btn btn-sm btn-info">üìä –°—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞</a>',
                obj.current_count,
                obj.target_count,
                percentage,
                reverse('admin:survey_statistics', args=[obj.id])
            )
        return '-'
    get_completion_info.short_description = _('–í—ã–ø–æ–ª–Ω–µ–Ω–∏–µ')
    get_completion_info.allow_tags = True
    
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('survey-stats/<int:task_id>/', 
                 self.admin_site.admin_view(self.survey_statistics_view), 
                 name='survey_statistics'),
        ]
        return custom_urls + urls
    
    def survey_statistics_view(self, request, task_id):
        """View for detailed survey statistics."""
        task = get_object_or_404(Task, id=task_id)
        
        # –û–±—â–∞—è —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞
        total_responses = SurveyAnswer.objects.filter(question__task=task).count()
        unique_clients = SurveyAnswer.objects.filter(question__task=task).values('client').distinct().count()
        
        # –°—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞ –ø–æ –≤–æ–ø—Ä–æ—Å–∞–º
        questions_stats = []
        for question in task.questions.all():
            question_stats = {
                'question': question,
                'total_answers': SurveyAnswer.objects.filter(question=question).count()
            }
            
            # –û–±—Ä–∞–±–æ—Ç–∫–∞ –≤—Å–µ—Ö —Ç–∏–ø–æ–≤ –≤–æ–ø—Ä–æ—Å–æ–≤ —Å –≤—ã–±–æ—Ä–æ–º
            if question.question_type in ['RADIO', 'CHECKBOX', 'SELECT_SINGLE', 'SELECT_MULTIPLE']:
                choice_stats = []
                
                # 1. –û–±—Ä–∞–±–æ—Ç–∫–∞ –∫–∞—Å—Ç–æ–º–Ω—ã—Ö –≤–∞—Ä–∏–∞–Ω—Ç–æ–≤ –æ—Ç–≤–µ—Ç–æ–≤ (–µ—Å–ª–∏ –µ—Å—Ç—å)
                if question.has_custom_choices():
                    for choice in question.choices.all():
                        # –ü–æ–¥—Å—á–µ—Ç –æ—Ç–≤–µ—Ç–æ–≤ –¥–ª—è —Ä–∞–∑–Ω—ã—Ö —Ç–∏–ø–æ–≤ –≤–æ–ø—Ä–æ—Å–æ–≤
                        if question.question_type == 'RADIO':
                            # –î–ª—è —Ä–∞–¥–∏–æ–∫–Ω–æ–ø–æ–∫ - —á–µ—Ä–µ–∑ selected_choices
                            count = SurveyAnswer.objects.filter(
                                question=question,
                                selected_choices=choice
                            ).count()
                        elif question.question_type == 'CHECKBOX':
                            # –î–ª—è —á–µ–∫–±–æ–∫—Å–æ–≤ - —á–µ—Ä–µ–∑ selected_choices
                            count = SurveyAnswer.objects.filter(
                                question=question,
                                selected_choices=choice
                            ).count()
                        elif question.question_type == 'SELECT_SINGLE':
                            # –î–ª—è –æ–¥–∏–Ω–æ—á–Ω–æ–≥–æ –≤—ã–±–æ—Ä–∞ - —á–µ—Ä–µ–∑ text_answer (ID –≤–∞—Ä–∏–∞–Ω—Ç–∞)
                            count = SurveyAnswer.objects.filter(
                                question=question,
                                text_answer=str(choice.id)
                            ).count()
                        elif question.question_type == 'SELECT_MULTIPLE':
                            # –î–ª—è –º–Ω–æ–∂–µ—Å—Ç–≤–µ–Ω–Ω–æ–≥–æ –≤—ã–±–æ—Ä–∞ - —á–µ—Ä–µ–∑ text_answer (ID –≤–∞—Ä–∏–∞–Ω—Ç–æ–≤ —á–µ—Ä–µ–∑ –∑–∞–ø—è—Ç—É—é)
                            count = SurveyAnswer.objects.filter(
                                question=question,
                                text_answer__contains=str(choice.id)
                            ).count()
                        
                        percentage = (count / question_stats['total_answers'] * 100) if question_stats['total_answers'] > 0 else 0
                        # –û–∫—Ä—É–≥–ª–µ–Ω–∏–µ –¥–æ 0.5%
                        percentage = round(percentage * 2) / 2.0
                        choice_stats.append({
                            'choice': choice,
                            'count': count,
                            'percentage': percentage
                        })
                
                # 2. –û–±—Ä–∞–±–æ—Ç–∫–∞ —Å—Ç–∞–Ω–¥–∞—Ä—Ç–Ω—ã—Ö –≤–∞—Ä–∏–∞–Ω—Ç–æ–≤ (–µ—Å–ª–∏ –Ω–µ—Ç –∫–∞—Å—Ç–æ–º–Ω—ã—Ö)
                else:
                    # –î–ª—è –≤–æ–ø—Ä–æ—Å–æ–≤ —Å —Ç–∏–ø–æ–º RADIO - —Å—Ç–∞–Ω–¥–∞—Ä—Ç–Ω—ã–µ –≤–∞—Ä–∏–∞–Ω—Ç—ã "–î–∞" –∏ "–ù–µ—Ç"
                    if question.question_type == 'RADIO':
                        # –°—Ç–∞–Ω–¥–∞—Ä—Ç–Ω—ã–µ –≤–∞—Ä–∏–∞–Ω—Ç—ã "–î–∞" –∏ "–ù–µ—Ç"
                        yes_count = SurveyAnswer.objects.filter(
                            question=question,
                            text_answer__iexact='–¥–∞'
                        ).count()
                        no_count = SurveyAnswer.objects.filter(
                            question=question,
                            text_answer__iexact='–Ω–µ—Ç'
                        ).count()
                        
                        yes_percentage = (yes_count / question_stats['total_answers'] * 100) if question_stats['total_answers'] > 0 else 0
                        no_percentage = (no_count / question_stats['total_answers'] * 100) if question_stats['total_answers'] > 0 else 0
                        # –û–∫—Ä—É–≥–ª–µ–Ω–∏–µ –¥–æ 0.5%
                        yes_percentage = round(yes_percentage * 2) / 2.0
                        no_percentage = round(no_percentage * 2) / 2.0
                        
                        choice_stats.extend([
                            {
                                'choice': type('Choice', (), {'choice_text': '–î–∞'}),
                                'count': yes_count,
                                'percentage': yes_percentage
                            },
                            {
                                'choice': type('Choice', (), {'choice_text': '–ù–µ—Ç'}),
                                'count': no_count,
                                'percentage': no_percentage
                            }
                        ])
                    
                    # –î–ª—è –≤–æ–ø—Ä–æ—Å–æ–≤ —Å —Ç–∏–ø–æ–º CHECKBOX - —Å—Ç–∞–Ω–¥–∞—Ä—Ç–Ω—ã–µ –≤–∞—Ä–∏–∞–Ω—Ç—ã "–î–∞" –∏ "–ù–µ—Ç"
                    elif question.question_type == 'CHECKBOX':
                        # –°—Ç–∞–Ω–¥–∞—Ä—Ç–Ω—ã–µ –≤–∞—Ä–∏–∞–Ω—Ç—ã "–î–∞" –∏ "–ù–µ—Ç"
                        yes_count = SurveyAnswer.objects.filter(
                            question=question,
                            text_answer__icontains='–¥–∞'
                        ).count()
                        no_count = SurveyAnswer.objects.filter(
                            question=question,
                            text_answer__icontains='–Ω–µ—Ç'
                        ).count()
                        
                        yes_percentage = (yes_count / question_stats['total_answers'] * 100) if question_stats['total_answers'] > 0 else 0
                        no_percentage = (no_count / question_stats['total_answers'] * 100) if question_stats['total_answers'] > 0 else 0
                        # –û–∫—Ä—É–≥–ª–µ–Ω–∏–µ –¥–æ 0.5%
                        yes_percentage = round(yes_percentage * 2) / 2.0
                        no_percentage = round(no_percentage * 2) / 2.0
                        
                        choice_stats.extend([
                            {
                                'choice': type('Choice', (), {'choice_text': '–î–∞'}),
                                'count': yes_count,
                                'percentage': yes_percentage
                            },
                            {
                                'choice': type('Choice', (), {'choice_text': '–ù–µ—Ç'}),
                                'count': no_count,
                                'percentage': no_percentage
                            }
                        ])
                
                question_stats['choice_stats'] = choice_stats
                
            # –¢–µ–∫—Å—Ç–æ–≤—ã–µ –≤–æ–ø—Ä–æ—Å—ã
            elif question.question_type in ['TEXT', 'TEXT_SHORT', 'TEXTAREA']:
                text_answers = SurveyAnswer.objects.filter(
                    question=question
                ).exclude(text_answer__isnull=True).exclude(text_answer='')
                question_stats['text_answers_count'] = text_answers.count()
                
            # –§–æ—Ç–æ –≤–æ–ø—Ä–æ—Å—ã
            elif question.question_type == 'PHOTO':
                # –ü–æ–ª—É—á–∞–µ–º –≤—Å–µ –æ—Ç–≤–µ—Ç—ã —Å —Ñ–æ—Ç–æ –¥–ª—è —ç—Ç–æ–≥–æ –≤–æ–ø—Ä–æ—Å–∞
                answers_with_photos = SurveyAnswer.objects.filter(
                    question=question
                ).prefetch_related('photos', 'client').order_by('client__name', 'created_at')
                
                # –ì—Ä—É–ø–ø–∏—Ä—É–µ–º —Ñ–æ—Ç–æ –ø–æ –æ—Ç–≤–µ—Ç–∞–º (–∫–ª–∏–µ–Ω—Ç–∞–º)
                photo_groups = []
                for answer in answers_with_photos:
                    photos_data = []
                    for photo in answer.photos.all():
                        # –ü–æ–ø—ã—Ç–∫–∞ –∏–∑–≤–ª–µ—á—å EXIF –¥–∞–Ω–Ω—ã–µ
                        exif_data = self._extract_photo_exif(photo.photo.path) if photo.photo else None
                        address = self._format_address_from_exif(exif_data) if exif_data else None
                        
                        photos_data.append({
                            'photo': photo,
                            'client': answer.client,
                            'created_at': answer.created_at,
                            'exif_data': exif_data,
                            'address': address
                        })
                    
                    if photos_data:
                        photo_groups.append({
                            'answer': answer,
                            'photos_data': photos_data
                        })
                
                question_stats['photo_groups'] = photo_groups
                question_stats['total_photos'] = sum(len(group['photos_data']) for group in photo_groups)
            
            questions_stats.append(question_stats)
        
        context = {
            'title': f'–°—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞: {task.title}',
            'task': task,
            'total_responses': total_responses,
            'unique_clients': unique_clients,
            'questions_stats': questions_stats,
            'opts': self.model._meta,
        }
        return render(request, 'admin/tasks/survey_statistics.html', context)

    def _extract_photo_exif(self, photo_path):
        """–ò–∑–≤–ª–µ–∫–∞–µ—Ç EXIF –¥–∞–Ω–Ω—ã–µ –∏–∑ —Ñ–æ—Ç–æ."""
        try:
            from PIL import Image
            from PIL.ExifTags import TAGS
            img = Image.open(photo_path)
            exifdata = img.getexif()
            if exifdata:
                exif = {}
                for tag_id in exifdata:
                    tag = TAGS.get(tag_id, tag_id)
                    data = exifdata.get(tag_id)
                    if isinstance(data, bytes):
                        data = data.decode()
                    exif[tag] = data
                return exif
        except Exception as e:
            print(f"Error extracting EXIF: {e}")
        return None

    def _format_address_from_exif(self, exif_data):
        """–§–æ—Ä–º–∞—Ç–∏—Ä—É–µ—Ç –∞–¥—Ä–µ—Å –∏–∑ EXIF –¥–∞–Ω–Ω—ã—Ö."""
        if not exif_data:
            return None
        
        try:
            # –ò–∑–≤–ª–µ—á–µ–Ω–∏–µ –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç –∏–∑ GPS –¥–∞–Ω–Ω—ã—Ö
            gps_info = exif_data.get('GPSInfo')
            if gps_info:
                # –ò–∑–≤–ª–µ—á–µ–Ω–∏–µ GPS-–∫–æ–æ—Ä–¥–∏–Ω–∞—Ç
                gps_keys = list(gps_info.keys())
                gps_values = list(gps_info.values())
                
                # –ò–∑–≤–ª–µ—á–µ–Ω–∏–µ —à–∏—Ä–æ—Ç—ã –∏ –¥–æ–ª–≥–æ—Ç—ã
                lat = self._convert_to_degrees(gps_info.get(2), gps_info.get(1))  # GPSLatitude, GPSLatitudeRef
                lon = self._convert_to_degrees(gps_info.get(4), gps_info.get(3))  # GPSLongitude, GPSLongitudeRef
                
                if lat and lon:
                    return f"{lat}, {lon}"
        except Exception as e:
            print(f"Error formatting address from EXIF: {e}")
        
        return None

    def _convert_to_degrees(self, value, ref):
        """–ö–æ–Ω–≤–µ—Ä—Ç–∏—Ä—É–µ—Ç GPS-–∫–æ–æ—Ä–¥–∏–Ω–∞—Ç—ã –≤ –≥—Ä–∞–¥—É—Å—ã."""
        if not value:
            return None
        
        try:
            d = float(value[0])
            m = float(value[1])
            s = float(value[2])
            
            degrees = d + (m / 60.0) + (s / 3600.0)
            
            if ref in ['S', 'W']:
                degrees = -degrees
                
            return degrees
        except (IndexError, ValueError, TypeError):
            return None

# –û—Å—Ç–∞–ª—å–Ω—ã–µ —Ä–µ–≥–∏—Å—Ç—Ä–∞—Ü–∏–∏ –º–æ–¥–µ–ª–µ–π...
class SurveyAnswerAdmin(admin.ModelAdmin):
    list_display = ('user', 'question', 'client', 'get_selected_choices', 'text_answer_preview', 'has_photos', 'created_at')
    readonly_fields = ('user', 'question', 'selected_choices', 'text_answer', 'client', 'created_at')
    list_per_page = 20
    change_list_template = 'admin/tasks/grouped_survey_answers.html'
    
    # Add filters
    list_filter = (
        'created_at',
        'user', 
        'question__task',
        'question',
        'client',
    )
    
    # Search fields for autocomplete functionality
    search_fields = (
        'client__name__icontains',  # For client search
        'user__username__icontains',  # For user search
        'user__first_name__icontains',  # For user first name
        'user__last_name__icontains',  # For user last name
        'question__task__title__icontains',  # For task search
        'question__question_text__icontains',  # For question search
    )
    
    def has_add_permission(self, request):
        return False
    def has_change_permission(self, request, obj=None):
        return False
    
    def get_selected_choices(self, obj):
        if obj.selected_choices.exists():
            return ', '.join([choice.choice_text for choice in obj.selected_choices.all()])
        return '-'
    get_selected_choices.short_description = _('–í—ã–±—Ä–∞–Ω–Ω—ã–µ –≤–∞—Ä–∏–∞–Ω—Ç—ã')
    
    def text_answer_preview(self, obj):
        if obj.text_answer:
            return obj.text_answer[:50] + '...' if len(obj.text_answer) > 50 else obj.text_answer
        return '-'
    text_answer_preview.short_description = _('–¢–µ–∫—Å—Ç–æ–≤—ã–π –æ—Ç–≤–µ—Ç')
    
    def has_photos(self, obj):
        return obj.photos.exists()
    has_photos.short_description = _('–ï—Å—Ç—å —Ñ–æ—Ç–æ')
    has_photos.boolean = True
    
    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs.select_related('user', 'question', 'client').prefetch_related('photos', 'selected_choices')

    def changelist_view(self, request, extra_context=None):
        """Override the default changelist view to use our grouped view"""
        from clients.models import Client
        from users.models import CustomUser
        from .models import Task
        
        clients = Client.objects.all()
        users = CustomUser.objects.filter(role='EMPLOYEE')
        moderators = CustomUser.objects.filter(role='MODERATOR')
        tasks = Task.objects.all()
        
        context = {
            'clients': clients,
            'users': users,
            'moderators': moderators,
            'tasks': tasks,
            'current_filters': request.GET,
            'opts': self.model._meta,
        }
        context.update(extra_context or {})
        
        return render(request, 'admin/tasks/grouped_survey_answers.html', context)

    def get_urls(self):
        urls = super().get_urls()
        # Remove the default changelist URL and replace with our custom functionality
        custom_urls = [
            path('export-excel/<int:task_id>/', 
                 self.admin_site.admin_view(self.export_excel_view), 
                 name='export_survey_answers_excel'),
            path('api/grouped-answers/', 
                 self.admin_site.admin_view(getGroupedAnswers), 
                 name='grouped_answers_api'),
            path('api/mark-as-read/', 
                 self.admin_site.admin_view(markAsRead), 
                 name='mark_as_read_api'),
            path('autocomplete_clients/', 
                 self.admin_site.admin_view(autocomplete_clients), 
                 name='autocomplete_clients'),
            path('autocomplete_tasks/', 
                 self.admin_site.admin_view(autocomplete_tasks), 
                 name='autocomplete_tasks'),
        ]
        return custom_urls + urls

    def export_excel_view(self, request, task_id):
        """Export survey answers for a specific task to Excel."""
        from .models import Task
        
        try:
            task = Task.objects.get(id=task_id)
        except Task.DoesNotExist:
            return HttpResponse("Task not found", status=404)
        
        # Get all answers for this task
        answers = SurveyAnswer.objects.filter(
            question__task=task
        ).select_related(
            'user', 'question', 'client'
        ).prefetch_related(
            'photos', 'selected_choices'
        ).order_by('client__name', 'question__order', 'created_at')
        
        # Create Excel workbook
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = f"–û—Ç–≤–µ—Ç—ã {task.title[:30]}"  # Limit sheet name length
        
        # Define headers
        headers = [
            '–ö–ª–∏–µ–Ω—Ç', '–°–æ—Ç—Ä—É–¥–Ω–∏–∫', '–î–∞—Ç–∞ –æ—Ç–≤–µ—Ç–∞', '–í–æ–ø—Ä–æ—Å', '–¢–∏–ø –≤–æ–ø—Ä–æ—Å–∞', 
            '–í—ã–±—Ä–∞–Ω–Ω—ã–µ –≤–∞—Ä–∏–∞–Ω—Ç—ã', '–¢–µ–∫—Å—Ç–æ–≤—ã–π –æ—Ç–≤–µ—Ç', '–ö–æ–ª–∏—á–µ—Å—Ç–≤–æ —Ñ–æ—Ç–æ'
        ]
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Write data
        row_num = 2
        for answer in answers:
            # Get selected choices as text
            selected_choices_text = ', '.join([choice.choice_text for choice in answer.selected_choices.all()])
            
            # Count photos
            photo_count = answer.photos.count()
            
            row_data = [
                answer.client.name,
                answer.user.get_full_name() or answer.user.username,
                answer.created_at.strftime('%d.%m.%Y %H:%M:%S'),
                answer.question.question_text,
                answer.question.get_question_type_display(),
                selected_choices_text,
                answer.text_answer,
                photo_count
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_num, column=col_num, value=str(value) if value is not None else '')
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            row_num += 1
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Limit width to 50
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Prepare response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="survey_answers_{task.title.replace(" ", "_")}_{timezone.now().strftime("%Y%m%d")}.xlsx"'
        
        workbook.save(response)
        return response

@admin.register(SurveyAnswerGroupReadStatus)
class SurveyAnswerGroupReadStatusAdmin(admin.ModelAdmin):
    list_display = ('task', 'client', 'user', 'date_created', 'read_at', 'read_by')
    list_filter = ('date_created', 'read_at', 'task', 'client', 'user')
    search_fields = ('task__title', 'client__name', 'user__username')
    readonly_fields = ('created_at',)
    list_per_page = 20


@admin.register(SurveyAnswer)
class SurveyAnswerAdminWrapper(SurveyAnswerAdmin):
    pass

# SurveyAnswerPhoto is hidden from admin as per requirements
# @admin.register(SurveyAnswerPhoto)
# class SurveyAnswerPhotoAdmin(admin.ModelAdmin):
#     list_display = ('answer', 'photo_thumbnail', 'created_at')
#     readonly_fields = ('answer', 'photo', 'created_at')
#     
#     def has_add_permission(self, request):
#         return False
#     
#     def photo_thumbnail(self, obj):
#         if obj.photo:
#             return format_html('<img src="{}" style="width: 50px; height: 50px; object-fit: cover;" />', obj.photo.url)
#         return '-'
#     photo_thumbnail.short_description = _('–ú–∏–Ω–∏–∞—Ç—é—Ä–∞')

@admin.register(PhotoReport)
class PhotoReportAdmin(admin.ModelAdmin):
    list_display = ('task', 'client', 'address', 'stand_count', 'created_by', 'created_at')
    readonly_fields = ('task', 'client', 'address', 'stand_count', 'comment', 'created_by')
    list_per_page = 20

# PhotoReportItem is hidden from admin as per requirements
# @admin.register(PhotoReportItem)
# class PhotoReportItemAdmin(admin.ModelAdmin):
#     list_display = ('report', 'photo_thumbnail', 'quality_score', 'is_accepted', 'created_at')
#     readonly_fields = ('report', 'photo', 'description', 'quality_score', 'is_accepted', 'created_at')
#     list_per_page = 20
#     
#     def photo_thumbnail(self, obj):
#         if obj.photo:
#             return format_html('<img src="{}" style="width: 50px; height: 50px; object-fit: cover;" />', obj.photo.url)
#         return '-'
#     photo_thumbnail.short_description = _('–ú–∏–Ω–∏–∞—Ç—é—Ä–∞')